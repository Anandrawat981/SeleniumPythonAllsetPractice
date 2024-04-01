import openpyxl
from openpyxl.styles import PatternFill

def getRowcount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumncount(file,sheetName):
    workbook= openpyxl.load_workbook(file)
    sheet =workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnum):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnum).value

def writeData(file,sheetName,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum,columnnum).value=data
    workbook.save(file)

def fillRedcolor(file,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfill = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rownum,columnnum).fill=redfill
    workbook.save(file)

def fillGreencolor(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenfill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rownum, columnnum).fill = greenfill
    workbook.save(file)