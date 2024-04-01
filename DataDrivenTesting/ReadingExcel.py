import openpyxl
import os

#FIle => workbook -->Sheets --->Rows-->Cells

file_path="C:/Users/Anand/Documents/data.xlsx"
#location=os.getcwd()+'/data1.xlsx'
workbook=openpyxl.load_workbook(file_path)
sheet=workbook["Sheet2"]
rows = sheet.max_row
columns = sheet.max_column

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='  ')
    print()