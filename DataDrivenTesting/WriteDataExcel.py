from random import random, randint

import openpyxl
file_name="C:/Users/Anand/Documents/WriteData.xlsx"
workbook=openpyxl.load_workbook(file_name)
sheet_name=workbook["data2"] #if we have multiple sheet then we are using this
#sheet_name=workbook.active #Here this will get the active sheet which is open and it is appliable for the single sheet
for r in range(1,6):
    for c in range(1,4):
        #sheet_name.cell(r,c).value="Rawat"
        sheet_name.cell(r,c).value=randint(1,10)
workbook.save(file_name)